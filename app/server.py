#!/usr/bin/env python3
import base64
import hashlib
import html
import json
import mimetypes
import os
import re
import secrets
import sqlite3
import sys
import zipfile
from collections import defaultdict
from datetime import datetime, timedelta
from http import cookies
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
APP_DIR = ROOT / "app"
CHUNK_SIZE = 1024 * 1024
SHEET_NAME_CLEANER = re.compile(r'[\\/*?:\[\]]')
DATE_NUMBER_FORMAT = "yyyy-mm-dd hh:mm"
FORMER_WORD = re.compile(r"\bformer\b", re.IGNORECASE)
VALID_SORT_KEYS = {"account_name", "updated_desc", "owner", "contract_status", "msa"}
VALID_DETAIL_TABS = {
    "timeline",
    "contacts",
    "deals",
    "cases",
    "notes",
    "attachments",
    "raw",
}
SESSION_COOKIE_NAME = os.environ.get("APP_SESSION_COOKIE_NAME", "cdxcrm_session")
SESSION_TTL_DAYS = max(1, int(os.environ.get("APP_SESSION_TTL_DAYS", "14")))
PBKDF2_ITERATIONS = 310_000

SERVICE_FEE_FIELDS = [
    ("MBS - Med A", "MBS - Med A", "currency"),
    ("MBS - Med B", "MBS - Med B", "currency"),
    ("MBS - Non-Medicare", "MBS - Non-Medicare", "currency"),
    ("MBS Discount for Multiple Studies", "MBS Discount for Multiple Studies", "text"),
    ("FEES", "FEES", "currency"),
    ("Endosheath Charge", "Endosheath Charge", "currency"),
    ("AmbuScope Charge", "AmbuScope Charge", "currency"),
    ("Hospital or DFW - MBS Flat Fee", "Hospital or DFW - MBS Flat Fee", "currency"),
    ("Hospital Flat Fee - 2 patients", "Hospital Flat Fee - 2 patients", "currency"),
    ("Hospital Flat Fee - 3 patients", "Hospital Flat Fee - 3 patients", "currency"),
    ("Hospital Flat Fee - 4+ patients", "Hospital Flat Fee - 4+ patients", "currency"),
    ("Travel Fee, if applicable", "Travel Fee, if applicable", "currency"),
    ("Billing Terms", "Billing Terms", "text"),
]


def configured_path(env_name, default):
    raw = os.environ.get(env_name, "").strip()
    if not raw:
        return default
    candidate = Path(raw).expanduser()
    if not candidate.is_absolute():
        candidate = (ROOT / candidate).resolve()
    return candidate


DB_PATH = configured_path("ZOHO_DB_PATH", ROOT / "data" / "zoho.sqlite3")
BACKUP_DIR = configured_path("ZOHO_BACKUP_DIR", ROOT / "zoho backup")
APP_STATE_DB_PATH = configured_path(
    "APP_STATE_DB_PATH", ROOT / "data" / "app_state.sqlite3"
)


def now_utc():
    return datetime.utcnow()


def utc_timestamp(dt=None):
    return (dt or now_utc()).strftime("%Y-%m-%dT%H:%M:%SZ")


def parse_utc_timestamp(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%dT%H:%M:%SZ")
    except ValueError:
        return None


def default_preferences():
    return {
        "theme": "system",
        "savedViews": [],
        "sortBy": "account_name",
        "defaultDetailTab": "timeline",
    }


def normalize_saved_view(view):
    if not isinstance(view, dict):
        return None

    name = clean_text(view.get("name"))[:80]
    if not name:
        return None
    view_id = clean_text(view.get("id"))[:120] or f"view_{secrets.token_hex(8)}"
    filters = view.get("filters") if isinstance(view.get("filters"), dict) else {}

    def string_list(key):
        values = filters.get(key)
        if not isinstance(values, list):
            return []
        cleaned = sorted(
            {
                clean_text(item)[:120]
                for item in values
                if clean_text(item)
            },
            key=str.casefold,
        )
        return cleaned[:100]

    sort_by = clean_text(filters.get("sortBy")) or "account_name"
    if sort_by not in VALID_SORT_KEYS:
        sort_by = "account_name"

    return {
        "id": view_id,
        "name": name,
        "filters": {
            "q": clean_text(filters.get("q"))[:200],
            "accountTypes": string_list("accountTypes"),
            "states": string_list("states"),
            "msas": string_list("msas"),
            "contractStatuses": string_list("contractStatuses"),
            "sortBy": sort_by,
        },
    }


def sanitize_preferences(payload):
    defaults = default_preferences()
    if not isinstance(payload, dict):
        return defaults

    preferences = {}
    theme = clean_text(payload.get("theme")) or defaults["theme"]
    preferences["theme"] = theme if theme in {"system", "light", "dark"} else defaults["theme"]

    sort_by = clean_text(payload.get("sortBy")) or defaults["sortBy"]
    preferences["sortBy"] = sort_by if sort_by in VALID_SORT_KEYS else defaults["sortBy"]

    default_tab = clean_text(payload.get("defaultDetailTab")) or defaults["defaultDetailTab"]
    preferences["defaultDetailTab"] = (
        default_tab if default_tab in VALID_DETAIL_TABS else defaults["defaultDetailTab"]
    )

    raw_views = payload.get("savedViews")
    normalized_views = []
    if isinstance(raw_views, list):
        for item in raw_views[:50]:
            view = normalize_saved_view(item)
            if view:
                normalized_views.append(view)
    preferences["savedViews"] = normalized_views
    return preferences


def connect_app_state():
    APP_STATE_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(APP_STATE_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_app_state_db():
    conn = connect_app_state()
    try:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE COLLATE NOCASE,
                display_name TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                is_admin INTEGER NOT NULL DEFAULT 0,
                must_change_password INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS user_preferences (
                user_id INTEGER PRIMARY KEY,
                preferences_json TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                token_hash TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL,
                last_seen_at TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );

            CREATE INDEX IF NOT EXISTS idx_sessions_token_hash ON sessions(token_hash);
            CREATE INDEX IF NOT EXISTS idx_sessions_expires_at ON sessions(expires_at);
            """
        )
        columns = {
            row["name"]
            for row in conn.execute("PRAGMA table_info(users)").fetchall()
        }
        if "is_admin" not in columns:
            conn.execute(
                "ALTER TABLE users ADD COLUMN is_admin INTEGER NOT NULL DEFAULT 0"
            )
        if "must_change_password" not in columns:
            conn.execute(
                "ALTER TABLE users ADD COLUMN must_change_password INTEGER NOT NULL DEFAULT 0"
            )
        conn.commit()
    finally:
        conn.close()


def hash_password(password, *, salt=None, iterations=PBKDF2_ITERATIONS):
    secret = password.encode("utf-8")
    salt_bytes = salt or secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", secret, salt_bytes, iterations)
    return "pbkdf2_sha256${}${}${}".format(
        iterations,
        base64.b64encode(salt_bytes).decode("ascii"),
        base64.b64encode(digest).decode("ascii"),
    )


def verify_password(password, stored_hash):
    try:
        algorithm, iterations, salt_b64, digest_b64 = stored_hash.split("$", 3)
    except ValueError:
        return False
    if algorithm != "pbkdf2_sha256":
        return False
    try:
        salt = base64.b64decode(salt_b64.encode("ascii"))
        expected = base64.b64decode(digest_b64.encode("ascii"))
        candidate = hashlib.pbkdf2_hmac(
            "sha256", password.encode("utf-8"), salt, int(iterations)
        )
    except (ValueError, TypeError):
        return False
    return secrets.compare_digest(candidate, expected)


def hash_session_token(token):
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def user_count():
    conn = connect_app_state()
    try:
        row = conn.execute("SELECT COUNT(*) AS count FROM users").fetchone()
        return row["count"] if row else 0
    finally:
        conn.close()


def list_local_users():
    conn = connect_app_state()
    try:
        rows = conn.execute(
            """
            SELECT username, display_name, is_active, is_admin, must_change_password, created_at, updated_at
            FROM users
            ORDER BY username COLLATE NOCASE
            """
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()


def upsert_local_user(
    username,
    password,
    *,
    display_name=None,
    is_active=True,
    is_admin=None,
    must_change_password=None,
):
    ensure_app_state_db()
    username = clean_text(username)
    if not username:
        raise ValueError("Username is required")
    if not password:
        raise ValueError("Password is required")

    display_name = clean_text(display_name) or username
    now = utc_timestamp()
    password_hash = hash_password(password)

    conn = connect_app_state()
    try:
        existing = conn.execute(
            "SELECT id, is_admin, must_change_password FROM users WHERE username = ? COLLATE NOCASE", (username,)
        ).fetchone()
        if existing:
            next_is_admin = existing["is_admin"] if is_admin is None else (1 if is_admin else 0)
            next_must_change = (
                existing["must_change_password"]
                if must_change_password is None
                else (1 if must_change_password else 0)
            )
            conn.execute(
                """
                UPDATE users
                SET display_name = ?, password_hash = ?, is_active = ?, is_admin = ?, must_change_password = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    display_name,
                    password_hash,
                    1 if is_active else 0,
                    next_is_admin,
                    next_must_change,
                    now,
                    existing["id"],
                ),
            )
            user_id = existing["id"]
        else:
            cursor = conn.execute(
                """
                INSERT INTO users (username, display_name, password_hash, is_active, is_admin, must_change_password, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    username,
                    display_name,
                    password_hash,
                    1 if is_active else 0,
                    1 if is_admin else 0,
                    1 if must_change_password else 0,
                    now,
                    now,
                ),
            )
            user_id = cursor.lastrowid

        existing_pref = conn.execute(
            "SELECT 1 FROM user_preferences WHERE user_id = ?", (user_id,)
        ).fetchone()
        if existing_pref is None:
            conn.execute(
                """
                INSERT INTO user_preferences (user_id, preferences_json, updated_at)
                VALUES (?, ?, ?)
                """,
                (user_id, json.dumps(default_preferences()), now),
            )
        conn.commit()
    finally:
        conn.close()


def revoke_user_sessions(user_id):
    conn = connect_app_state()
    try:
        conn.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))
        conn.commit()
    finally:
        conn.close()


def set_local_user_password(username, password, *, must_change_password=False):
    if not password:
        raise ValueError("Password is required")
    ensure_app_state_db()
    conn = connect_app_state()
    try:
        row = conn.execute(
            "SELECT id FROM users WHERE username = ? COLLATE NOCASE", (username,)
        ).fetchone()
        if row is None:
            raise ValueError(f"User not found: {username}")
        conn.execute(
            "UPDATE users SET password_hash = ?, must_change_password = ?, updated_at = ? WHERE id = ?",
            (
                hash_password(password),
                1 if must_change_password else 0,
                utc_timestamp(),
                row["id"],
            ),
        )
        conn.commit()
        user_id = row["id"]
    finally:
        conn.close()
    revoke_user_sessions(user_id)


def set_local_user_admin(username, is_admin=True):
    ensure_app_state_db()
    conn = connect_app_state()
    try:
        row = conn.execute(
            "SELECT id FROM users WHERE username = ? COLLATE NOCASE", (username,)
        ).fetchone()
        if row is None:
            raise ValueError(f"User not found: {username}")
        conn.execute(
            "UPDATE users SET is_admin = ?, updated_at = ? WHERE id = ?",
            (1 if is_admin else 0, utc_timestamp(), row["id"]),
        )
        conn.commit()
    finally:
        conn.close()


def user_id_by_username(username):
    conn = connect_app_state()
    try:
        row = conn.execute(
            "SELECT id FROM users WHERE username = ? COLLATE NOCASE", (username,)
        ).fetchone()
        return row["id"] if row else None
    finally:
        conn.close()


def delete_session_by_token(token):
    if not token:
        return
    conn = connect_app_state()
    try:
        conn.execute("DELETE FROM sessions WHERE token_hash = ?", (hash_session_token(token),))
        conn.commit()
    finally:
        conn.close()


def create_session(user_id):
    ensure_app_state_db()
    token = secrets.token_urlsafe(32)
    now = now_utc()
    stamp = utc_timestamp(now)
    expires_at = utc_timestamp(now + timedelta(days=SESSION_TTL_DAYS))
    conn = connect_app_state()
    try:
        conn.execute(
            """
            INSERT INTO sessions (user_id, token_hash, created_at, last_seen_at, expires_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (user_id, hash_session_token(token), stamp, stamp, expires_at),
        )
        conn.commit()
    finally:
        conn.close()
    return token


def build_session_cookie(token):
    parts = [
        f"{SESSION_COOKIE_NAME}={token}",
        "Path=/",
        "HttpOnly",
        "SameSite=Lax",
        f"Max-Age={SESSION_TTL_DAYS * 24 * 60 * 60}",
    ]
    if os.environ.get("APP_SESSION_COOKIE_SECURE", "").strip().lower() in {
        "1",
        "true",
        "yes",
    }:
        parts.append("Secure")
    return "; ".join(parts)


def clear_session_cookie():
    parts = [
        f"{SESSION_COOKIE_NAME}=",
        "Path=/",
        "HttpOnly",
        "SameSite=Lax",
        "Max-Age=0",
    ]
    if os.environ.get("APP_SESSION_COOKIE_SECURE", "").strip().lower() in {
        "1",
        "true",
        "yes",
    }:
        parts.append("Secure")
    return "; ".join(parts)


def load_preferences_for_user(user_id):
    conn = connect_app_state()
    try:
        row = conn.execute(
            "SELECT preferences_json FROM user_preferences WHERE user_id = ?", (user_id,)
        ).fetchone()
        if row is None:
            return default_preferences()
        try:
            payload = json.loads(row["preferences_json"])
        except json.JSONDecodeError:
            payload = {}
        return sanitize_preferences(payload)
    finally:
        conn.close()


def save_preferences_for_user(user_id, preferences):
    sanitized = sanitize_preferences(preferences)
    conn = connect_app_state()
    try:
        conn.execute(
            """
            INSERT INTO user_preferences (user_id, preferences_json, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(user_id) DO UPDATE SET
                preferences_json = excluded.preferences_json,
                updated_at = excluded.updated_at
            """,
            (user_id, json.dumps(sanitized), utc_timestamp()),
        )
        conn.commit()
    finally:
        conn.close()
    return sanitized


def authenticate_local_user(username, password):
    ensure_app_state_db()
    conn = connect_app_state()
    try:
        row = conn.execute(
            """
            SELECT id, username, display_name, password_hash, is_active, is_admin, must_change_password
            FROM users
            WHERE username = ? COLLATE NOCASE
            """,
            (clean_text(username),),
        ).fetchone()
        if row is None or not row["is_active"]:
            return None
        if not verify_password(password, row["password_hash"]):
            return None
        return {
            "id": row["id"],
            "username": row["username"],
            "displayName": row["display_name"],
            "isAdmin": bool(row["is_admin"]),
            "mustChangePassword": bool(row["must_change_password"]),
        }
    finally:
        conn.close()


def change_password_for_user(user_id, new_password, *, clear_force_flag=True):
    if not new_password:
        raise ValueError("New password is required")
    conn = connect_app_state()
    try:
        updates = ["password_hash = ?", "updated_at = ?"]
        values = [hash_password(new_password), utc_timestamp()]
        if clear_force_flag:
            updates.append("must_change_password = 0")
        values.append(user_id)
        conn.execute(
            f"UPDATE users SET {', '.join(updates)} WHERE id = ?",
            values,
        )
        conn.commit()
    finally:
        conn.close()


def verify_current_password(user_id, password):
    conn = connect_app_state()
    try:
        row = conn.execute(
            "SELECT password_hash FROM users WHERE id = ?", (user_id,)
        ).fetchone()
        if row is None:
            return False
        return verify_password(password, row["password_hash"])
    finally:
        conn.close()


def change_own_password(user, current_password, new_password):
    if user.get("mustChangePassword"):
        change_password_for_user(user["id"], new_password, clear_force_flag=True)
        return
    if not verify_current_password(user["id"], current_password):
        raise ValueError("Current password is incorrect")
    change_password_for_user(user["id"], new_password, clear_force_flag=True)


def admin_reset_password(actor, target_username, new_password):
    if not actor.get("isAdmin"):
        raise PermissionError("Administrator access required")
    target_username = clean_text(target_username)
    if not target_username:
        raise ValueError("Target username is required")
    if target_username.lower() == str(actor.get("username", "")).lower():
        raise ValueError("Use the personal change-password form for your own account")
    conn = connect_app_state()
    try:
        row = conn.execute(
            """
            SELECT id, username, display_name, is_admin
            FROM users
            WHERE username = ? COLLATE NOCASE
            """,
            (target_username,),
        ).fetchone()
        if row is None:
            raise ValueError("User not found")
        conn.execute(
            """
            UPDATE users
            SET password_hash = ?, must_change_password = 1, updated_at = ?
            WHERE id = ?
            """,
            (hash_password(new_password), utc_timestamp(), row["id"]),
        )
        conn.commit()
        target = {
            "id": row["id"],
            "username": row["username"],
            "displayName": row["display_name"],
            "isAdmin": bool(row["is_admin"]),
        }
    finally:
        conn.close()
    revoke_user_sessions(target["id"])
    return target


def current_user_from_session(token):
    if not token:
        return None
    ensure_app_state_db()
    token_hash = hash_session_token(token)
    conn = connect_app_state()
    try:
        conn.execute("DELETE FROM sessions WHERE expires_at <= ?", (utc_timestamp(),))
        row = conn.execute(
            """
            SELECT u.id, u.username, u.display_name, u.is_active, u.is_admin, u.must_change_password, s.id AS session_id
            FROM sessions s
            JOIN users u ON u.id = s.user_id
            WHERE s.token_hash = ? AND s.expires_at > ?
            """,
            (token_hash, utc_timestamp()),
        ).fetchone()
        if row is None or not row["is_active"]:
            conn.execute("DELETE FROM sessions WHERE token_hash = ?", (token_hash,))
            conn.commit()
            return None
        conn.execute(
            "UPDATE sessions SET last_seen_at = ? WHERE id = ?",
            (utc_timestamp(), row["session_id"]),
        )
        conn.commit()
        preferences = load_preferences_for_user(row["id"])
        return {
            "id": row["id"],
            "username": row["username"],
            "displayName": row["display_name"],
            "isAdmin": bool(row["is_admin"]),
            "mustChangePassword": bool(row["must_change_password"]),
            "preferences": preferences,
        }
    finally:
        conn.close()


def bootstrap_payload(user):
    return {
        "user": {
            "id": user["id"],
            "username": user["username"],
            "displayName": user["displayName"],
            "isAdmin": bool(user.get("isAdmin")),
            "mustChangePassword": bool(user.get("mustChangePassword")),
        },
        "preferences": sanitize_preferences(user.get("preferences")),
    }


def render_login_page(error_message=""):
    template = (APP_DIR / "login.html").read_text(encoding="utf-8")
    error_html = ""
    if error_message:
        error_html = (
            '<div class="auth-error" role="alert">'
            f"{html.escape(error_message)}"
            "</div>"
        )
    return template.replace("__LOGIN_ERROR__", error_html)


def render_change_password_page(user, error_message=""):
    template = (APP_DIR / "change_password.html").read_text(encoding="utf-8")
    error_html = ""
    if error_message:
        error_html = (
            '<div class="auth-error" role="alert">'
            f"{html.escape(error_message)}"
            "</div>"
        )
    return (
        template.replace("__DISPLAY_NAME__", html.escape(user["displayName"]))
        .replace("__CHANGE_PASSWORD_ERROR__", error_html)
    )


def render_index_page(user):
    template = (APP_DIR / "index.html").read_text(encoding="utf-8")
    bootstrap_json = json.dumps(bootstrap_payload(user), ensure_ascii=False).replace(
        "</", "<\\/"
    )
    return template.replace(
        "__BOOTSTRAP_JSON__",
        bootstrap_json,
    )


def q(identifier):
    return '"' + identifier.replace('"', '""') + '"'


def connect():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def table_exists(conn, table):
    return (
        conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type IN ('table', 'view') AND name = ?",
            (table,),
        ).fetchone()
        is not None
    )


def parse_int(value, default, minimum=1, maximum=500):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return max(minimum, min(maximum, parsed))


def chunked(values, size=400):
    for index in range(0, len(values), size):
        yield values[index : index + size]


def placeholders(values):
    return ", ".join("?" for _ in values)


def format_timestamp(dt=None):
    return (dt or datetime.now()).strftime("%Y-%m-%d_%H%M")


def parse_excel_datetime(value):
    if not value:
        return None
    text = str(value).strip()
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def clean_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    return text


def has_former_flag(value):
    return bool(FORMER_WORD.search(clean_text(value)))


def compact_account(row):
    return {
        "id": row["record_id"],
        "name": row["account_name"],
        "type": row["account_type"],
        "facilityType": row["facility_type"],
        "city": row["billing_city"],
        "state": row["billing_state"],
        "county": row["billing_county"],
        "msa": row["msa"],
        "territories": row["territories"],
        "contractStatus": row["contract_status"],
        "contractType": row["contract_type"],
        "products": row["products_available"],
        "owner": row["account_owner"],
        "beds": row["beds"],
        "modifiedTime": row["modified_time"],
    }


def fetch_kpis(conn):
    def one(sql, params=()):
        return conn.execute(sql, params).fetchone()[0]

    return {
        "accounts": one("SELECT COUNT(*) FROM accounts"),
        "contacts": one("SELECT COUNT(*) FROM contacts"),
        "deals": one("SELECT COUNT(*) FROM deals"),
        "cases": one("SELECT COUNT(*) FROM cases"),
        "accountNotes": one("SELECT COUNT(*) FROM notes_accounts"),
        "contactNotes": one("SELECT COUNT(*) FROM notes_contacts"),
        "msaMappings": one("SELECT COUNT(*) FROM msa_mappings"),
    }


def filtered_values(params, key):
    return [value.strip() for value in params.get(key, []) if value.strip()]


def account_search_filters(params, exclude_field=None):
    term = params.get("q", [""])[0].strip()
    account_types = filtered_values(params, "account_type")
    states = filtered_values(params, "state")
    msas = filtered_values(params, "msa")
    contract_statuses = filtered_values(params, "contract_status")
    where = []
    values = []

    if term:
        like = f"%{term}%"
        where.append(
            """
            (
                account_name LIKE ?
                OR billing_city LIKE ?
                OR billing_state LIKE ?
                OR msa LIKE ?
                OR territories LIKE ?
                OR account_owner LIKE ?
                OR products_available LIKE ?
                OR contract_status LIKE ?
            )
            """
        )
        values.extend([like] * 8)

    multi_filters = [
        ("account_type", "account_type", account_types),
        ("state", "billing_state", states),
        ("msa", "msa", msas),
        ("contract_status", "contract_status", contract_statuses),
    ]
    for field_name, column, selections in multi_filters:
        if field_name == exclude_field or not selections:
            continue
        where.append(f"{column} IN ({', '.join('?' for _ in selections)})")
        values.extend(selections)

    return where, values


def account_order_by(params):
    sort_key = params.get("sort", ["account_name"])[0].strip() or "account_name"
    options = {
        "account_name": "account_name COLLATE NOCASE, record_id",
        "updated_desc": "modified_time DESC, account_name COLLATE NOCASE, record_id",
        "owner": "account_owner COLLATE NOCASE, account_name COLLATE NOCASE, record_id",
        "contract_status": "contract_status COLLATE NOCASE, account_name COLLATE NOCASE, record_id",
        "msa": "msa COLLATE NOCASE, account_name COLLATE NOCASE, record_id",
    }
    return options.get(sort_key, options["account_name"])


def build_account_query(
    params,
    *,
    select_sql="*",
    from_sql="account_search",
    group_by=None,
    order_by=None,
    limit=None,
    offset=None,
    exclude_field=None,
    extra_where=None,
    extra_values=None,
):
    # Shared query builder so the UI pages and the Excel export stay in lockstep
    # on filtering semantics while still selecting different columns/shapes.
    where, values = account_search_filters(params, exclude_field=exclude_field)
    if extra_where:
        where.extend(extra_where)
    if extra_values:
        values.extend(extra_values)

    where_sql = f"WHERE {' AND '.join(where)}" if where else ""
    sql = f"SELECT {select_sql} FROM {from_sql} {where_sql}"
    if group_by:
        sql += f" GROUP BY {group_by}"
    if order_by is None:
        order_by = account_order_by(params)
    if order_by:
        sql += f" ORDER BY {order_by}"
    if limit is not None:
        sql += " LIMIT ?"
        values.append(limit)
        if offset is not None:
            sql += " OFFSET ?"
            values.append(offset)
    elif offset is not None:
        sql += " LIMIT -1 OFFSET ?"
        values.append(offset)
    return sql, values


def search_accounts(params):
    limit = parse_int(params.get("limit", ["50"])[0], 50, 1, 100)
    offset = parse_int(params.get("offset", ["0"])[0], 0, 0, 1000000)
    conn = connect()
    try:
        total_sql, total_values = build_account_query(
            params,
            select_sql="COUNT(*)",
            from_sql="account_search",
            order_by="",
        )
        rows_sql, row_values = build_account_query(
            params,
            from_sql="account_search",
            limit=limit,
            offset=offset,
        )
        total = conn.execute(total_sql, total_values).fetchone()[0]
        rows = conn.execute(rows_sql, row_values).fetchall()
        return {
            "total": total,
            "limit": limit,
            "offset": offset,
            "accounts": [compact_account(row) for row in rows],
        }
    finally:
        conn.close()


def list_filter_values(field, params):
    allowed = {
        "account_type": "account_type",
        "state": "billing_state",
        "msa": "msa",
        "contract_status": "contract_status",
    }
    column = allowed.get(field)
    if not column:
        return {"values": []}
    conn = connect()
    try:
        sql, values = build_account_query(
            params,
            select_sql=f"{q(column)} AS value, COUNT(*) AS count",
            from_sql="account_search",
            group_by=q(column),
            order_by="count DESC, value COLLATE NOCASE",
            limit=250,
            exclude_field=field,
            extra_where=[f"COALESCE({q(column)}, '') <> ''"],
        )
        rows = conn.execute(sql, values).fetchall()
        return {
            "values": [{"value": row["value"], "count": row["count"]} for row in rows]
        }
    finally:
        conn.close()


def fetch_raw_account(conn, account_id):
    row = conn.execute(
        "SELECT * FROM accounts WHERE record_id = ?", (account_id,)
    ).fetchone()
    if not row:
        return None
    raw = json.loads(row["raw_data"])
    return raw


def child_rows(conn, table, where_column, value, fields, limit=100):
    selected = ", ".join(q(field) for field in fields)
    rows = conn.execute(
        f"""
        SELECT {selected}
        FROM {q(table)}
        WHERE {q(where_column)} = ?
        ORDER BY COALESCE(modified_time, created_time) DESC
        LIMIT ?
        """,
        (value, limit),
    ).fetchall()
    return [dict(row) for row in rows]


def notes(conn, table, account_id, limit=100):
    rows = conn.execute(
        f"""
        SELECT record_id, created_time, modified_time, note_title, note_content, note_owner
        FROM {q(table)}
        WHERE parent_id_id = ?
        ORDER BY COALESCE(modified_time, created_time) DESC
        LIMIT ?
        """,
        (account_id, limit),
    ).fetchall()
    return [dict(row) for row in rows]


def account_notes(conn, account_id, limit=100):
    if table_exists(conn, "notes_unified"):
        rows = conn.execute(
            """
            SELECT
                record_id,
                parent_module,
                created_time,
                modified_time,
                note_title,
                note_content_preview AS note_content,
                record_status,
                '' AS note_owner
            FROM notes_unified
            WHERE rollup_account_id = ?
            ORDER BY COALESCE(modified_time, created_time) DESC
            LIMIT ?
            """,
            (account_id, limit),
        ).fetchall()
        return [dict(row) for row in rows]
    return notes(conn, "notes_accounts", account_id, limit=limit)


def timeline_event(
    event_type,
    event_id,
    at,
    title,
    summary="",
    *,
    detail="",
    badge="",
    status="",
    owner="",
    source_id="",
    source_tab="",
):
    return {
        "id": event_id,
        "type": event_type,
        "at": at or "",
        "title": title or "",
        "summary": summary or "",
        "detail": detail or "",
        "badge": badge or "",
        "status": status or "",
        "owner": owner or "",
        "sourceId": source_id or "",
        "sourceTab": source_tab or "",
    }


def build_account_timeline(account, related):
    events = [
        timeline_event(
            "account",
            f"account:{account['id']}",
            account.get("modifiedTime") or account.get("createdTime"),
            "Account updated",
            "Facility record",
            detail=account.get("owner") or "",
            status=account.get("contractStatus") or "",
            source_id=account.get("id") or "",
        )
    ]

    for note in related["notes"]:
        events.append(
            timeline_event(
                "note",
                f"note:{note['record_id']}",
                note.get("modified_time") or note.get("created_time"),
                note.get("note_title") or "Untitled note",
                note.get("note_content") or "",
                badge=note.get("parent_module") or "Account note",
                owner=note.get("note_owner") or "",
                source_id=note["record_id"],
                source_tab="notes",
            )
        )

    for deal in related["deals"]:
        summary_parts = [
            deal.get("stage"),
            deal.get("contact_name"),
            deal.get("amount"),
        ]
        events.append(
            timeline_event(
                "deal",
                f"deal:{deal['record_id']}",
                deal.get("modified_time") or deal.get("closing_date"),
                deal.get("deal_name") or "Deal update",
                " · ".join(part for part in summary_parts if part),
                detail=deal.get("type") or "",
                status=deal.get("stage") or "",
                source_id=deal["record_id"],
                source_tab="deals",
            )
        )

    for case in related["cases"]:
        summary_parts = [
            case.get("case_number"),
            case.get("case_origin"),
            case.get("priority"),
        ]
        events.append(
            timeline_event(
                "case",
                f"case:{case['record_id']}",
                case.get("modified_time") or case.get("created_time"),
                case.get("subject") or "Case update",
                " · ".join(part for part in summary_parts if part),
                detail=case.get("description") or "",
                status=case.get("status") or "",
                source_id=case["record_id"],
                source_tab="cases",
            )
        )

    for contact in related["contacts"]:
        summary_parts = [
            contact.get("title") or contact.get("contact_type") or contact.get("role"),
            contact.get("email"),
            contact.get("phone") or contact.get("mobile"),
        ]
        events.append(
            timeline_event(
                "contact",
                f"contact:{contact['record_id']}",
                contact.get("modified_time") or contact.get("created_time"),
                contact.get("contact_name") or "Contact updated",
                " · ".join(part for part in summary_parts if part),
                status=contact.get("contact_type") or "",
                source_id=contact["record_id"],
                source_tab="contacts",
            )
        )

    for email in related["emails"]:
        summary_parts = [
            email.get("sender"),
            email.get("sent_to"),
        ]
        detail_parts = []
        if email.get("has_attachment"):
            count = email.get("attachment_count") or "1"
            detail_parts.append(f"{count} attachment(s)")
        if email.get("attachment_name"):
            detail_parts.append(email.get("attachment_name"))
        events.append(
            timeline_event(
                "email",
                f"email:{email['record_id']}",
                email.get("sent_on") or email.get("modified_time") or email.get("created_time"),
                email.get("subject") or "Email activity",
                " -> ".join(part for part in summary_parts if part),
                detail=" · ".join(part for part in detail_parts if part),
                badge=email.get("module") or "",
                status=email.get("status") or "",
            )
        )

    for meeting in related["meetings"]:
        summary_parts = [
            meeting.get("meeting_type"),
            meeting.get("location"),
        ]
        detail = " to ".join(
            part for part in [meeting.get("from_time"), meeting.get("to_time")] if part
        )
        events.append(
            timeline_event(
                "meeting",
                f"meeting:{meeting['record_id']}",
                meeting.get("from_time") or meeting.get("modified_time") or meeting.get("created_time"),
                meeting.get("title") or "Meeting scheduled",
                " · ".join(part for part in summary_parts if part),
                detail=detail,
                badge=meeting.get("related_to_module") or "",
            )
        )

    for task in related["tasks"]:
        summary_parts = [
            task.get("priority"),
            task.get("due_date"),
        ]
        events.append(
            timeline_event(
                "task",
                f"task:{task['record_id']}",
                task.get("modified_time") or task.get("created_time") or task.get("due_date"),
                task.get("subject") or "Task updated",
                " · ".join(part for part in summary_parts if part),
                badge=task.get("related_to_module") or "",
                status=task.get("status") or "",
            )
        )

    events.sort(
        key=lambda event: (
            event["at"] or "",
            event["title"].lower(),
            event["id"],
        ),
        reverse=True,
    )
    return events


def service_fees(raw):
    return [
        {"label": label, "value": raw.get(source, ""), "kind": kind}
        for label, source, kind in SERVICE_FEE_FIELDS
    ]


def file_extension(filename):
    return Path(filename or "").suffix.lower()


def attachment_size(row):
    return (
        row["zip_file_size"]
        if row["zip_file_size"] is not None
        else row["metadata_file_size"]
    )


def compact_attachment(row, availability):
    return {
        "attachment_record_id": row["attachment_record_id"],
        "original_filename": row["original_filename"],
        "parent_module": row["parent_module"],
        "parent_record_id": row["parent_record_id"],
        "rollup_account_id": row["rollup_account_id"],
        "file_extension": file_extension(row["original_filename"]),
        "file_size": attachment_size(row),
        "created_time": row["created_time"],
        "modified_time": row["modified_time"],
        "record_status": row["record_status"],
        "mapping_confidence": row["mapping_confidence"],
        "zip_file": row["zip_file"],
        "zip_inner_path": row["zip_inner_path"],
        "availability": availability,
        "download_url": f"/api/attachments/{quote(row['attachment_record_id'], safe='')}/download",
    }


def attachment_availability(rows):
    statuses = {}
    by_zip = {}
    for row in rows:
        attachment_id = row["attachment_record_id"]
        zip_file = row["zip_file"]
        inner_path = row["zip_inner_path"]
        if not zip_file or not inner_path:
            statuses[attachment_id] = "unmatched"
            continue
        by_zip.setdefault(zip_file, []).append((attachment_id, inner_path))

    for zip_file, entries in by_zip.items():
        zip_path = BACKUP_DIR / zip_file
        if not zip_path.exists():
            for attachment_id, _ in entries:
                statuses[attachment_id] = "missing_zip"
            continue
        try:
            with zipfile.ZipFile(zip_path) as archive:
                names = set(archive.namelist())
        except zipfile.BadZipFile:
            for attachment_id, _ in entries:
                statuses[attachment_id] = "bad_zip"
            continue
        for attachment_id, inner_path in entries:
            statuses[attachment_id] = (
                "available" if inner_path in names else "missing_zip_entry"
            )
    return statuses


def account_attachment_sources(conn):
    parts = [
        "SELECT 'Accounts' AS parent_module, ? AS parent_record_id, ? AS rollup_account_id"
    ]
    values = []
    for table, module, account_column in (
        ("contacts", "Contacts", "account_name_id"),
        ("deals", "Deals", "account_name_id"),
        ("cases", "Cases", "account_name_id"),
        ("notes_unified", "Notes", "rollup_account_id"),
        ("referrals_c", "Referrals_C", "rollup_account_id"),
        ("zohosign_documents_c", "ZohoSign Documents_C", "rollup_account_id"),
        ("emails", "Emails", "rollup_account_id"),
        ("meetings", "Meetings", "rollup_account_id"),
        ("tasks", "Tasks", "rollup_account_id"),
    ):
        if table_exists(conn, table):
            parts.append(
                f"""
                SELECT ? AS parent_module, record_id AS parent_record_id, {q(account_column)} AS rollup_account_id
                FROM {q(table)}
                WHERE {q(account_column)} = ?
                """
            )
            values.append(module)
    return parts, values


def account_attachments(account_id):
    conn = connect()
    try:
        if not table_exists(conn, "attachments"):
            return {
                "accountId": account_id,
                "total": 0,
                "groups": [],
                "attachments": [],
            }

        source_parts, modules = account_attachment_sources(conn)
        values = [account_id, account_id]
        for module in modules:
            values.extend([module, account_id])

        rows = conn.execute(
            f"""
            WITH account_related(parent_module, parent_record_id, rollup_account_id) AS (
                {" UNION ALL ".join(source_parts)}
            )
            SELECT DISTINCT
                attachment.attachment_record_id,
                attachment.parent_record_id,
                attachment.parent_module,
                attachment.original_filename,
                attachment.zip_file,
                attachment.zip_inner_path,
                attachment.metadata_file_size,
                attachment.zip_file_size,
                attachment.created_time,
                attachment.modified_time,
                attachment.record_status,
                attachment.mapping_confidence,
                account_related.rollup_account_id
            FROM attachments attachment
            JOIN account_related
              ON attachment.parent_module = account_related.parent_module
             AND attachment.parent_record_id = account_related.parent_record_id
            ORDER BY
                attachment.parent_module COLLATE NOCASE,
                COALESCE(attachment.modified_time, attachment.created_time) DESC,
                attachment.original_filename COLLATE NOCASE
            """,
            values,
        ).fetchall()

        statuses = attachment_availability(rows)
        attachments = [
            compact_attachment(
                row, statuses.get(row["attachment_record_id"], "unknown")
            )
            for row in rows
        ]
        counts = {}
        for attachment in attachments:
            module = attachment["parent_module"] or "Unknown"
            counts[module] = counts.get(module, 0) + 1
        return {
            "accountId": account_id,
            "total": len(attachments),
            "groups": [
                {"module": module, "count": counts[module]} for module in sorted(counts)
            ],
            "attachments": attachments,
        }
    finally:
        conn.close()


def fetch_attachment_record(attachment_record_id):
    conn = connect()
    try:
        if not table_exists(conn, "attachments"):
            return None
        return conn.execute(
            """
            SELECT attachment_record_id, original_filename, zip_file, zip_inner_path, zip_file_size, metadata_file_size
            FROM attachments
            WHERE attachment_record_id = ?
            """,
            (attachment_record_id,),
        ).fetchone()
    finally:
        conn.close()


def account_detail(account_id):
    conn = connect()
    try:
        raw = fetch_raw_account(conn, account_id)
        if raw is None:
            return None

        account = {
            "id": raw.get("Record Id"),
            "name": raw.get("Account Name"),
            "type": raw.get("Account Type"),
            "facilityType": raw.get("Facility Type"),
            "phone": raw.get("Phone"),
            "website": raw.get("Website"),
            "address": {
                "street": raw.get("Billing Street"),
                "street2": raw.get("Billing Street 2"),
                "city": raw.get("Billing City"),
                "state": raw.get("Billing State"),
                "zip": raw.get("Billing Code"),
                "county": raw.get("Billing County"),
                "country": raw.get("Billing Country"),
            },
            "owner": raw.get("Account Owner"),
            "msa": raw.get("MSA"),
            "territories": raw.get("Territories"),
            "contractStatus": raw.get("Contract Status"),
            "contractType": raw.get("Contract Type"),
            "productsAvailable": raw.get("Products Available"),
            "productsContracted": raw.get("Products Contracted"),
            "totalBeds": raw.get("Total Beds"),
            "certifiedMedicareBeds": raw.get("Certified Medicare Beds"),
            "facilityNpi": raw.get("Facility NPI"),
            "chainAccountId": raw.get("Chain Account.id"),
            "chainAccount": raw.get("Chain Account"),
            "networkParentAccountId": raw.get("Network Parent Account.id"),
            "networkParentAccount": raw.get("Network Parent Account"),
            "billingPhone": raw.get("Billing Phone"),
            "billingEmail": raw.get("Billing Email"),
            "createdTime": raw.get("Created Time"),
            "modifiedTime": raw.get("Modified Time"),
            "description": raw.get("Description"),
            "serviceFees": service_fees(raw),
            "raw": raw,
        }

        contacts = child_rows(
            conn,
            "contacts",
            "account_name_id",
            account_id,
            [
                "record_id",
                "contact_name",
                "first_name",
                "last_name",
                "title",
                "email",
                "phone",
                "mobile",
                "contact_type",
                "role",
                "modified_time",
            ],
        )
        deals = child_rows(
            conn,
            "deals",
            "account_name_id",
            account_id,
            [
                "record_id",
                "deal_name",
                "stage",
                "type",
                "amount",
                "closing_date",
                "probability_pct",
                "contact_name",
                "modified_time",
            ],
        )
        cases = child_rows(
            conn,
            "cases",
            "account_name_id",
            account_id,
            [
                "record_id",
                "case_number",
                "status",
                "priority",
                "case_origin",
                "subject",
                "description",
                "modified_time",
            ],
        )
        note_rows = account_notes(conn, account_id)
        emails = child_rows(
            conn,
            "emails",
            "rollup_account_id",
            account_id,
            [
                "record_id",
                "subject",
                "module",
                "sender",
                "sent_to",
                "sent_on",
                "status",
                "has_attachment",
                "attachment_count",
                "attachment_name",
                "created_time",
                "modified_time",
            ],
        ) if table_exists(conn, "emails") else []
        meetings = child_rows(
            conn,
            "meetings",
            "rollup_account_id",
            account_id,
            [
                "record_id",
                "title",
                "from_time",
                "to_time",
                "location",
                "meeting_type",
                "related_to_module",
                "created_time",
                "modified_time",
            ],
        ) if table_exists(conn, "meetings") else []
        tasks = child_rows(
            conn,
            "tasks",
            "rollup_account_id",
            account_id,
            [
                "record_id",
                "subject",
                "status",
                "priority",
                "due_date",
                "related_to_module",
                "created_time",
                "modified_time",
            ],
        ) if table_exists(conn, "tasks") else []

        related = {
            "contacts": contacts,
            "deals": deals,
            "cases": cases,
            "notes": note_rows,
            "emails": emails,
            "meetings": meetings,
            "tasks": tasks,
        }

        return {
            "account": account,
            "related": related,
            "timeline": build_account_timeline(account, related),
        }
    finally:
        conn.close()


def table_columns(conn, table):
    return {row["name"] for row in conn.execute(f"PRAGMA table_info({q(table)})")}


def build_address(row):
    return ", ".join(
        part
        for part in [row.get("billing_street"), row.get("billing_street_2")]
        if part
    )


def account_stat_rows(conn, table, account_column, account_ids):
    if not account_ids or not table_exists(conn, table):
        return []

    rows = []
    for batch in chunked(account_ids):
        sql = f"""
            SELECT
                {q(account_column)} AS account_id,
                COUNT(*) AS item_count,
                MAX(COALESCE(modified_time, created_time)) AS latest_date
            FROM {q(table)}
            WHERE {q(account_column)} IN ({placeholders(batch)})
            GROUP BY {q(account_column)}
        """
        rows.extend(conn.execute(sql, batch).fetchall())
    return rows


def contact_inventory_rows(conn, account_ids):
    if not account_ids or not table_exists(conn, "contacts"):
        return []

    rows = []
    for batch in chunked(account_ids):
        sql = f"""
            SELECT
                account.record_id AS linked_account_id,
                account.account_name AS linked_account_name,
                account.msa AS msa,
                contact.record_id,
                contact.contact_name,
                contact.first_name,
                contact.last_name,
                contact.title,
                contact.contact_type,
                contact.role,
                contact.email,
                contact.phone,
                contact.mobile,
                contact.fax,
                contact.mailing_street,
                contact.mailing_city,
                contact.mailing_state,
                contact.mailing_zip,
                contact.mailing_country,
                contact.enrich_status
            FROM contacts contact
            JOIN accounts account ON account.record_id = contact.account_name_id
            WHERE contact.account_name_id IN ({placeholders(batch)})
            ORDER BY
                account.account_name COLLATE NOCASE,
                contact.contact_name COLLATE NOCASE,
                contact.record_id
        """
        rows.extend(conn.execute(sql, batch).fetchall())
    return [dict(row) for row in rows]


def attachment_rollup_sql(conn):
    parts = [
        "SELECT record_id AS account_id, 'Accounts' AS parent_module, record_id AS parent_record_id FROM accounts"
    ]
    for table, module, account_column in (
        ("contacts", "Contacts", "account_name_id"),
        ("deals", "Deals", "account_name_id"),
        ("cases", "Cases", "account_name_id"),
        ("notes_unified", "Notes", "rollup_account_id"),
        ("referrals_c", "Referrals_C", "rollup_account_id"),
        ("zohosign_documents_c", "ZohoSign Documents_C", "rollup_account_id"),
        ("emails", "Emails", "rollup_account_id"),
        ("meetings", "Meetings", "rollup_account_id"),
        ("tasks", "Tasks", "rollup_account_id"),
    ):
        if table_exists(conn, table):
            parts.append(
                f"""
                SELECT {q(account_column)} AS account_id, '{module}' AS parent_module, record_id AS parent_record_id
                FROM {q(table)}
                WHERE COALESCE({q(account_column)}, '') <> ''
                """
            )

    if not table_exists(conn, "notes_unified") and table_exists(conn, "notes_accounts"):
        parts.append(
            """
            SELECT parent_id_id AS account_id, 'Notes' AS parent_module, record_id AS parent_record_id
            FROM notes_accounts
            WHERE COALESCE(parent_id_id, '') <> ''
            """
        )
    return " UNION ALL ".join(parts)


def attachment_inventory_rows(conn, account_ids):
    # Attachment export intentionally stays metadata-only so a future ZIP export
    # can reuse this manifest without changing workbook structure or memory use.
    if not account_ids or not table_exists(conn, "attachments"):
        return []

    rollup_sql = attachment_rollup_sql(conn)
    rows = []
    for batch in chunked(account_ids):
        sql = f"""
            WITH account_related(account_id, parent_module, parent_record_id) AS (
                {rollup_sql}
            )
            SELECT DISTINCT
                account_related.account_id,
                account.account_name,
                account.msa,
                attachment.attachment_record_id,
                attachment.parent_record_id,
                attachment.parent_module,
                attachment.original_filename,
                attachment.zip_file,
                attachment.zip_inner_path,
                attachment.metadata_file_size,
                attachment.zip_file_size,
                attachment.created_time,
                attachment.modified_time,
                attachment.record_status,
                attachment.mapping_confidence,
                attachment.raw_data,
                attachment.indexed_at
            FROM attachments attachment
            JOIN account_related
              ON attachment.parent_module = account_related.parent_module
             AND attachment.parent_record_id = account_related.parent_record_id
            JOIN accounts account ON account.record_id = account_related.account_id
            WHERE account_related.account_id IN ({placeholders(batch)})
            ORDER BY
                account.account_name COLLATE NOCASE,
                attachment.parent_module COLLATE NOCASE,
                COALESCE(attachment.modified_time, attachment.created_time) DESC,
                attachment.original_filename COLLATE NOCASE
        """
        rows.extend(conn.execute(sql, batch).fetchall())

    inventory = []
    for row in rows:
        raw = json.loads(row["raw_data"]) if row["raw_data"] else {}
        filename = row["original_filename"] or ""
        inventory.append(
            {
                "account_id": row["account_id"],
                "account_name": row["account_name"],
                "msa": row["msa"] or "",
                "parent_module": row["parent_module"] or "",
                "parent_record_id": row["parent_record_id"] or "",
                "attachment_id": row["attachment_record_id"],
                "old_attachment_id": raw.get("Old Attachment Id", ""),
                "original_filename": filename,
                "file_extension": file_extension(filename),
                "mime_type": mimetypes.guess_type(filename)[0] or "",
                "file_size": attachment_size(row),
                "metadata_file_size": row["metadata_file_size"],
                "zip_file_size": row["zip_file_size"],
                "created_time": row["created_time"],
                "modified_time": row["modified_time"],
                "record_status": row["record_status"],
                "mapping_confidence": row["mapping_confidence"],
                "source_zip_file": row["zip_file"] or "",
                "source_zip_path": row["zip_inner_path"] or "",
                "internal_attachment_path": row["zip_inner_path"] or "",
                "documents_flag": raw.get("Documents", ""),
                "link_url": raw.get("Link URL", ""),
                "field_id": raw.get("Field.id", ""),
                "indexed_at": row["indexed_at"],
            }
        )
    return inventory


def contact_priority(contact):
    # Primary contact selection prefers explicit "main/primary" hints, then
    # scheduling-oriented roles, then contacts with email, and finally any
    # remaining linked contact so each facility row surfaces the most useful
    # operational person without duplicating facility rows.
    score = 0
    haystack = " ".join(
        str(contact.get(key, "") or "").lower()
        for key in ("role", "title", "contact_type")
    )
    if "main contact" in haystack or "primary" in haystack:
        score += 100
    if any(
        token in haystack
        for token in ("sched", "admission", "intake", "referral", "liaison", "coordinator")
    ):
        score += 40
    if contact.get("email"):
        score += 15
    if contact.get("phone") or contact.get("mobile"):
        score += 10
    if contact.get("title"):
        score += 5
    name = contact.get("contact_name") or ""
    return (
        -score,
        not bool(contact.get("email")),
        not bool(contact.get("phone") or contact.get("mobile")),
        name.lower(),
        contact.get("record_id", ""),
    )


def contact_rollups(contact_rows):
    grouped = defaultdict(list)
    for row in contact_rows:
        grouped[row["linked_account_id"]].append(row)

    rollups = {}
    for account_id, contacts in grouped.items():
        ordered = sorted(contacts, key=contact_priority)
        primary = ordered[0]
        rollups[account_id] = {
            "primary_contact": primary.get("contact_name", ""),
            "primary_contact_title": primary.get("title")
            or primary.get("contact_type")
            or primary.get("role", ""),
            "primary_contact_phone": primary.get("phone") or primary.get("mobile", ""),
            "primary_contact_email": primary.get("email", ""),
            "additional_contacts_count": max(len(ordered) - 1, 0),
        }
    return rollups


def enriched_account_rows(conn, params, *, include_attachments=True):
    sql, values = build_account_query(params, from_sql="accounts")
    account_rows = [dict(row) for row in conn.execute(sql, values).fetchall()]
    account_ids = [row["record_id"] for row in account_rows]

    referral_stats = {
        row["account_id"]: dict(row)
        for row in account_stat_rows(conn, "referrals_c", "rollup_account_id", account_ids)
    }
    contact_rows = contact_inventory_rows(conn, account_ids)
    contact_rollup_data = contact_rollups(contact_rows)
    attachments = attachment_inventory_rows(conn, account_ids) if include_attachments else []

    for row in account_rows:
        account_id = row["record_id"]
        row["address"] = build_address(row)
        row["referral_count"] = referral_stats.get(account_id, {}).get("item_count", 0)
        row.update(
            contact_rollup_data.get(
                account_id,
                {
                    "primary_contact": "",
                    "primary_contact_title": "",
                    "primary_contact_phone": "",
                    "primary_contact_email": "",
                    "additional_contacts_count": 0,
                },
            )
        )
    return account_rows, contact_rows, attachments


def account_export_columns(available):
    available = set(available)
    available.update(
        {
            "address",
            "referral_count",
            "primary_contact",
            "primary_contact_title",
            "primary_contact_phone",
            "primary_contact_email",
            "additional_contacts_count",
        }
    )
    columns = []
    for label, key, width, kind in (
        ("Account Name", "account_name", 36, "text"),
        ("Referral Count", "referral_count", 14, "number"),
        ("MSA / Market", "msa", 24, "text"),
        ("Status", "contract_status", 18, "text"),
        ("Facility Type", "facility_type", 18, "text"),
        ("Primary Contact", "primary_contact", 28, "text"),
        ("Primary Contact Title", "primary_contact_title", 24, "text"),
        ("Primary Contact Phone", "primary_contact_phone", 18, "text"),
        ("Primary Contact Email", "primary_contact_email", 28, "text"),
        ("Additional Contacts Count", "additional_contacts_count", 18, "number"),
        ("Address", "address", 28, "text"),
        ("City", "billing_city", 18, "text"),
        ("State", "billing_state", 10, "text"),
        ("ZIP", "billing_code", 12, "text"),
        ("Phone", "phone", 16, "text"),
        ("Website", "website", 24, "text"),
        ("Fees", "fees", 14, "text"),
        ("MBS - Med A", "mbs_med_a", 14, "text"),
        ("MBS - Med B", "mbs_med_b", 14, "text"),
        ("MBS - Non-Medicare", "mbs_non_medicare", 18, "text"),
        ("Hospital / DFW Flat Fee", "hospital_or_dfw_mbs_flat_fee", 20, "text"),
        ("Hospital Flat Fee - 2", "hospital_flat_fee_2_patients", 18, "text"),
        ("Hospital Flat Fee - 3", "hospital_flat_fee_3_patients", 18, "text"),
        ("Hospital Flat Fee - 4+", "hospital_flat_fee_4_patients", 18, "text"),
        ("Travel Fee", "travel_fee_if_applicable", 14, "text"),
        ("Endosheath Charge", "endosheath_charge", 18, "text"),
        ("MBS Discount", "mbs_discount_for_multiple_studies", 18, "text"),
        ("Account Type", "account_type", 18, "text"),
        ("Products Available", "products_available", 18, "text"),
        ("Products Contracted", "products_contracted", 18, "text"),
        ("Total Beds", "total_beds", 12, "text"),
        ("Certified Medicare Beds", "certified_medicare_beds", 18, "text"),
        ("Billing Phone", "billing_phone", 16, "text"),
        ("Billing Email", "billing_email", 24, "text"),
        ("Contract Type", "contract_type", 18, "text"),
        ("Account Owner", "account_owner", 20, "text"),
    ):
        if key in available:
            columns.append({"label": label, "key": key, "width": width, "kind": kind})
    return columns


def contacts_export_columns():
    return [
        {"label": "Account Name", "key": "linked_account_name", "width": 36, "kind": "text"},
        {"label": "MSA / Market", "key": "msa", "width": 24, "kind": "text"},
        {"label": "Contact Name", "key": "contact_name", "width": 28, "kind": "text"},
        {"label": "First Name", "key": "first_name", "width": 18, "kind": "text"},
        {"label": "Last Name", "key": "last_name", "width": 18, "kind": "text"},
        {"label": "Title", "key": "title", "width": 24, "kind": "text"},
        {"label": "Department", "key": "contact_type", "width": 20, "kind": "text"},
        {"label": "Contact Type / Role", "key": "contact_role_display", "width": 24, "kind": "text"},
        {"label": "Email", "key": "email", "width": 28, "kind": "text"},
        {"label": "Phone", "key": "phone", "width": 16, "kind": "text"},
        {"label": "Mobile Phone", "key": "mobile", "width": 16, "kind": "text"},
        {"label": "Fax", "key": "fax", "width": 16, "kind": "text"},
        {"label": "Address", "key": "mailing_street", "width": 28, "kind": "text"},
        {"label": "City", "key": "mailing_city", "width": 18, "kind": "text"},
        {"label": "State", "key": "mailing_state", "width": 10, "kind": "text"},
        {"label": "ZIP", "key": "mailing_zip", "width": 12, "kind": "text"},
        {"label": "Contact Status", "key": "contact_status", "width": 16, "kind": "text"},
        {"label": "Linked Account ID", "key": "linked_account_id", "width": 24, "kind": "text"},
        {"label": "Linked Account Name", "key": "linked_account_name", "width": 36, "kind": "text"},
    ]


def filtered_contacts_export_columns():
    return [
        {"label": "Site Name", "key": "site_name", "width": 36, "kind": "text"},
        {"label": "Contact Name", "key": "contact_name", "width": 28, "kind": "text"},
        {"label": "Title", "key": "title", "width": 24, "kind": "text"},
        {"label": "Type", "key": "type", "width": 20, "kind": "text"},
        {"label": "Email", "key": "email", "width": 28, "kind": "text"},
        {"label": "Phone", "key": "phone", "width": 16, "kind": "text"},
        {"label": "Cell / Mobile", "key": "cell_mobile", "width": 16, "kind": "text"},
    ]


def attachment_export_columns():
    return [
        {"label": "Account ID", "key": "account_id", "width": 24, "kind": "text"},
        {"label": "Account Name", "key": "account_name", "width": 32, "kind": "text"},
        {"label": "MSA", "key": "msa", "width": 24, "kind": "text"},
        {"label": "Parent Module", "key": "parent_module", "width": 18, "kind": "text"},
        {"label": "Parent Record ID", "key": "parent_record_id", "width": 24, "kind": "text"},
        {"label": "Attachment ID", "key": "attachment_id", "width": 30, "kind": "text"},
        {"label": "Old Attachment ID", "key": "old_attachment_id", "width": 24, "kind": "text"},
        {"label": "Original Filename", "key": "original_filename", "width": 36, "kind": "text"},
        {"label": "File Extension", "key": "file_extension", "width": 12, "kind": "text"},
        {"label": "File Type / MIME Type", "key": "mime_type", "width": 20, "kind": "text"},
        {"label": "File Size", "key": "file_size", "width": 14, "kind": "number"},
        {"label": "Metadata File Size", "key": "metadata_file_size", "width": 16, "kind": "number"},
        {"label": "ZIP File Size", "key": "zip_file_size", "width": 14, "kind": "number"},
        {"label": "Created Date", "key": "created_time", "width": 18, "kind": "date"},
        {"label": "Modified Date", "key": "modified_time", "width": 18, "kind": "date"},
        {"label": "Record Status", "key": "record_status", "width": 16, "kind": "text"},
        {"label": "Mapping Confidence", "key": "mapping_confidence", "width": 18, "kind": "text"},
        {"label": "Source ZIP File", "key": "source_zip_file", "width": 20, "kind": "text"},
        {"label": "Source ZIP Path", "key": "source_zip_path", "width": 36, "kind": "text"},
        {"label": "Internal Attachment Path", "key": "internal_attachment_path", "width": 36, "kind": "text"},
        {"label": "Documents Flag", "key": "documents_flag", "width": 14, "kind": "text"},
        {"label": "Link URL", "key": "link_url", "width": 28, "kind": "text"},
        {"label": "Field ID", "key": "field_id", "width": 18, "kind": "text"},
        {"label": "Indexed At", "key": "indexed_at", "width": 18, "kind": "date"},
    ]


def sheet_name_for_msa(msa, used_names):
    # MSA worksheet names need Excel-safe sanitization plus duplicate handling
    # because multiple markets can collapse to the same 31-char workbook title.
    base = SHEET_NAME_CLEANER.sub(" ", (msa or "").strip())
    base = " ".join(base.split()) or "Unknown MSA"
    base = base[:31]
    if base not in used_names:
        used_names.add(base)
        return base

    index = 2
    while True:
        suffix = f"_{index}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        index += 1


def write_sheet(ws, columns, rows):
    ws.append([column["label"] for column in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    widths = [column["width"] for column in columns]
    for row in rows:
        values = []
        for index, column in enumerate(columns):
            value = row.get(column["key"], "")
            if value is None:
                value = ""
            if column["kind"] == "date":
                parsed = parse_excel_datetime(value)
                value = parsed or value
            values.append(value)
            widths[index] = min(max(widths[index], len(str(value)) + 2), 48)
        ws.append(values)

    last_row = max(ws.max_row, 1)
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    for col_index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell, column in zip(row, columns):
            if column["kind"] == "date" and isinstance(cell.value, datetime):
                cell.number_format = DATE_NUMBER_FORMAT


def contacts_export_rows(contact_rows):
    # Contacts worksheet generation stays one row per contact so non-technical
    # reviewers can see the full facility contact inventory alongside the
    # one-row-per-facility operational tabs.
    rows = []
    for row in contact_rows:
        role_parts = [row.get("contact_type", ""), row.get("role", "")]
        rows.append(
            {
                **row,
                "contact_role_display": " / ".join(part for part in role_parts if part),
                "contact_status": row.get("enrich_status", ""),
            }
        )
    return rows


def qualifies_contact_for_export(contact):
    if has_former_flag(contact.get("title")) or has_former_flag(contact.get("contact_type")):
        return False
    return any(
        clean_text(contact.get(key))
        for key in ("email", "phone", "mobile")
    )


def filtered_contacts_export_rows(account_rows, contact_rows):
    accounts_by_id = {row["record_id"]: row for row in account_rows}
    qualifying_contacts_by_account = defaultdict(list)

    for contact in contact_rows:
        if not qualifies_contact_for_export(contact):
            continue
        account = accounts_by_id.get(contact.get("linked_account_id"))
        if account is None:
            continue
        qualifying_contacts_by_account[account["record_id"]].append(
            {
                "site_name": clean_text(account.get("account_name")),
                "contact_name": clean_text(contact.get("contact_name")),
                "title": clean_text(contact.get("title")),
                "type": clean_text(contact.get("contact_type")),
                "email": clean_text(contact.get("email")),
                "phone": clean_text(contact.get("phone")),
                "cell_mobile": clean_text(contact.get("mobile")),
                "contact_first_name_sort": clean_text(contact.get("first_name")),
                "contact_last_name_sort": clean_text(contact.get("last_name")),
                "contact_name_sort": clean_text(contact.get("contact_name")),
                "zoho_contact_id_sort": clean_text(contact.get("record_id")),
            }
        )

    return sorted(
        [row for rows in qualifying_contacts_by_account.values() for row in rows],
        key=lambda row: (
            row["site_name"].lower(),
            row["contact_last_name_sort"].lower(),
            row["contact_first_name_sort"].lower(),
            row["contact_name_sort"].lower(),
            row["zoho_contact_id_sort"].lower(),
        ),
    )


def workbook_bytes(params):
    conn = connect()
    try:
        # Workbook generation is intentionally data-first so the same collected
        # rows can later feed an Excel+attachments ZIP export mode.
        account_rows, contact_rows, attachment_rows = enriched_account_rows(conn, params)
        account_columns = account_export_columns(table_columns(conn, "accounts"))
        contacts_columns = contacts_export_columns()
        attachment_columns = attachment_export_columns()

        workbook = Workbook()
        workbook.remove(workbook.active)

        all_results = workbook.create_sheet("All Results")
        write_sheet(all_results, account_columns, account_rows)

        # MSA sheets are created from the already-filtered result set so each
        # workbook tab reflects the same search context as the main export tab.
        by_msa = defaultdict(list)
        for row in account_rows:
            by_msa[row.get("msa") or ""].append(row)

        used_sheet_names = {"All Results", "Contacts", "Attachments"}
        for msa in sorted(by_msa, key=lambda value: ((value or "").strip().lower() or "unknown msa")):
            ws = workbook.create_sheet(sheet_name_for_msa(msa, used_sheet_names))
            write_sheet(ws, account_columns, by_msa[msa])

        contacts_ws = workbook.create_sheet("Contacts")
        write_sheet(contacts_ws, contacts_columns, contacts_export_rows(contact_rows))

        attachments_ws = workbook.create_sheet("Attachments")
        write_sheet(attachments_ws, attachment_columns, attachment_rows)

        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()
    finally:
        conn.close()


def contacts_workbook_bytes(params):
    conn = connect()
    try:
        account_rows, contact_rows, _attachment_rows = enriched_account_rows(
            conn, params, include_attachments=False
        )
        export_rows = filtered_contacts_export_rows(account_rows, contact_rows)
        contacts_columns = filtered_contacts_export_columns()

        workbook = Workbook()
        workbook.remove(workbook.active)

        contacts_ws = workbook.create_sheet("Contacts")
        write_sheet(contacts_ws, contacts_columns, export_rows)

        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()
    finally:
        conn.close()


def metadata():
    conn = connect()
    try:
        rows = conn.execute(
            """
            SELECT table_name, zip_file, csv_file, row_count, column_count, imported_at
            FROM import_metadata
            ORDER BY table_name
            """
        ).fetchall()
        return {"kpis": fetch_kpis(conn), "imports": [dict(row) for row in rows]}
    finally:
        conn.close()


class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        print(f"{self.address_string()} - {format % args}")

    def send_json(self, payload, status=200, extra_headers=None):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        for key, value in extra_headers or []:
            self.send_header(key, value)
        self.end_headers()
        self.wfile.write(body)

    def send_file(self, path):
        if not path.exists() or not path.is_file():
            self.send_error(404)
            return
        content = path.read_bytes()
        mime = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", mime)
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def send_html(self, content, status=200, extra_headers=None):
        body = content.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        for key, value in extra_headers or []:
            self.send_header(key, value)
        self.end_headers()
        self.wfile.write(body)

    def send_redirect(self, location, extra_headers=None):
        self.send_response(302)
        self.send_header("Location", location)
        for key, value in extra_headers or []:
            self.send_header(key, value)
        self.end_headers()

    def send_bytes(self, content, content_type, filename):
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header(
            "Content-Disposition", f"attachment; filename*=UTF-8''{quote(filename)}"
        )
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def cookie_value(self, name):
        jar = cookies.SimpleCookie()
        jar.load(self.headers.get("Cookie", ""))
        morsel = jar.get(name)
        return morsel.value if morsel else ""

    def current_user(self):
        if hasattr(self, "_cached_user"):
            return self._cached_user
        token = self.cookie_value(SESSION_COOKIE_NAME)
        self._cached_user = current_user_from_session(token)
        return self._cached_user

    def require_auth(self, *, api=False, export=False, allow_password_change_only=False):
        user = self.current_user()
        if user is not None:
            if user.get("mustChangePassword") and not allow_password_change_only:
                if api:
                    self.send_json({"error": "Password change required"}, 423)
                else:
                    self.send_redirect("/change-password")
                return None
            return user
        if api:
            self.send_json({"error": "Authentication required"}, 401)
        elif export:
            self.send_redirect("/login")
        else:
            self.send_redirect("/login")
        return None

    def read_body(self):
        length = int(self.headers.get("Content-Length", "0") or 0)
        return self.rfile.read(length) if length else b""

    def read_json_body(self):
        raw = self.read_body()
        if not raw:
            return {}
        return json.loads(raw.decode("utf-8"))

    def read_form_body(self):
        raw = self.read_body().decode("utf-8")
        return parse_qs(raw, keep_blank_values=True)

    def stream_attachment(self, attachment_record_id, disposition="attachment"):
        row = fetch_attachment_record(attachment_record_id)
        if row is None:
            self.send_json({"error": "Attachment not found"}, 404)
            return
        if not row["zip_file"] or not row["zip_inner_path"]:
            self.send_json({"error": "Attachment has no matched zip entry"}, 404)
            return

        zip_path = BACKUP_DIR / row["zip_file"]
        if not zip_path.exists():
            self.send_json(
                {"error": f"Attachment zip not found: {row['zip_file']}"}, 404
            )
            return

        try:
            archive = zipfile.ZipFile(zip_path)
        except zipfile.BadZipFile:
            self.send_json(
                {"error": f"Attachment zip is unreadable: {row['zip_file']}"}, 500
            )
            return

        try:
            info = archive.getinfo(row["zip_inner_path"])
        except KeyError:
            archive.close()
            self.send_json({"error": "Attachment entry not found in zip"}, 404)
            return

        filename = row["original_filename"] or Path(row["zip_inner_path"]).name
        mime = mimetypes.guess_type(filename)[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-Type", mime)
        self.send_header("Content-Length", str(info.file_size))
        if disposition not in {"attachment", "inline"}:
            disposition = "attachment"
        self.send_header(
            "Content-Disposition", f"{disposition}; filename*=UTF-8''{quote(filename)}"
        )
        self.end_headers()

        try:
            with archive.open(info) as source:
                while True:
                    chunk = source.read(CHUNK_SIZE)
                    if not chunk:
                        break
                    self.wfile.write(chunk)
        finally:
            archive.close()

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        params = parse_qs(parsed.query)

        if path == "/login":
            user = self.current_user()
            if user is not None:
                self.send_redirect(
                    "/change-password" if user.get("mustChangePassword") else "/"
                )
                return
            self.send_html(render_login_page(params.get("error", [""])[0]))
            return

        if path == "/change-password":
            user = self.require_auth(allow_password_change_only=True)
            if user is None:
                return
            if not user.get("mustChangePassword"):
                self.send_redirect("/")
                return
            self.send_html(render_change_password_page(user, params.get("error", [""])[0]))
            return

        if path == "/":
            user = self.require_auth()
            if user is None:
                return
            self.send_html(render_index_page(user))
            return

        if path.startswith("/api/") or path.startswith("/export/"):
            if self.require_auth(api=path.startswith("/api/"), export=path.startswith("/export/")) is None:
                return

        if not DB_PATH.exists() and (
            path.startswith("/api/") or path.startswith("/export/")
        ):
            self.send_json(
                {"error": "Database not found. Run scripts/import_zoho.py first."}, 503
            )
            return

        try:
            if path == "/api/metadata":
                self.send_json(metadata())
                return
            if path == "/api/me":
                user = self.current_user()
                self.send_json(bootstrap_payload(user) if user else {"error": "Authentication required"}, 200 if user else 401)
                return
            if path == "/api/admin/users":
                user = self.require_auth(api=True)
                if user is None:
                    return
                if not user.get("isAdmin"):
                    self.send_json({"error": "Administrator access required"}, 403)
                    return
                self.send_json({"users": list_local_users()})
                return
            if path == "/api/accounts":
                self.send_json(search_accounts(params))
                return
            if path == "/api/filter-values":
                self.send_json(list_filter_values(params.get("field", [""])[0], params))
                return
            if path == "/export/accounts.xlsx":
                filename = f"cdx_zoho_accounts_export_{format_timestamp()}.xlsx"
                self.send_bytes(
                    workbook_bytes(params),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename,
                )
                return
            if path == "/export/contacts.xlsx":
                filename = f"cdx_zoho_contacts_export_{format_timestamp()}.xlsx"
                self.send_bytes(
                    contacts_workbook_bytes(params),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename,
                )
                return
            if path.startswith("/api/attachments/") and path.endswith("/download"):
                attachment_id = unquote(
                    path.removeprefix("/api/attachments/").removesuffix("/download")
                )
                self.stream_attachment(
                    attachment_id, params.get("disposition", ["attachment"])[0]
                )
                return
            if path.startswith("/api/accounts/") and path.endswith("/attachments"):
                account_id = unquote(
                    path.removeprefix("/api/accounts/").removesuffix("/attachments")
                )
                self.send_json(account_attachments(account_id))
                return
            if path.startswith("/api/accounts/"):
                account_id = path.rsplit("/", 1)[-1]
                payload = account_detail(account_id)
                if payload is None:
                    self.send_json({"error": "Account not found"}, 404)
                else:
                    self.send_json(payload)
                return
        except sqlite3.Error as exc:
            self.send_json({"error": str(exc)}, 500)
            return

        requested = (APP_DIR / path.lstrip("/")).resolve()
        if APP_DIR in requested.parents:
            self.send_file(requested)
        else:
            self.send_error(403)

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path

        if path == "/login":
            try:
                form = self.read_form_body()
                username = clean_text(form.get("username", [""])[0])
                password = form.get("password", [""])[0]
                user = authenticate_local_user(username, password)
            except (json.JSONDecodeError, UnicodeDecodeError, sqlite3.Error):
                user = None
            if user is None:
                message = "Invalid username or password"
                if user_count() == 0:
                    message = "No local users exist yet. Run scripts/manage_users.py seed first."
                self.send_redirect(f"/login?error={quote(message)}")
                return

            token = create_session(user["id"])
            self.send_redirect(
                "/change-password" if user.get("mustChangePassword") else "/",
                extra_headers=[("Set-Cookie", build_session_cookie(token))],
            )
            return

        if path == "/change-password":
            user = self.require_auth(allow_password_change_only=True)
            if user is None:
                return
            try:
                form = self.read_form_body()
                new_password = form.get("new_password", [""])[0]
                confirm_password = form.get("confirm_password", [""])[0]
                if not new_password:
                    raise ValueError("New password is required")
                if new_password != confirm_password:
                    raise ValueError("New passwords did not match")
                change_own_password(user, "", new_password)
            except ValueError as exc:
                self.send_redirect(f"/change-password?error={quote(str(exc))}")
                return
            self.send_redirect("/")
            return

        if path == "/api/logout":
            token = self.cookie_value(SESSION_COOKIE_NAME)
            if token:
                delete_session_by_token(token)
            self.send_json(
                {"ok": True},
                extra_headers=[("Set-Cookie", clear_session_cookie())],
            )
            return

        if path == "/api/preferences":
            user = self.require_auth(api=True)
            if user is None:
                return
            try:
                payload = self.read_json_body()
                preferences = save_preferences_for_user(user["id"], payload)
                self.send_json({"preferences": preferences})
            except (json.JSONDecodeError, UnicodeDecodeError):
                self.send_json({"error": "Invalid JSON body"}, 400)
            except sqlite3.Error as exc:
                self.send_json({"error": str(exc)}, 500)
            return

        if path == "/api/change-password":
            user = self.require_auth(api=True, allow_password_change_only=True)
            if user is None:
                return
            try:
                payload = self.read_json_body()
                current_password = payload.get("currentPassword", "")
                new_password = payload.get("newPassword", "")
                confirm_password = payload.get("confirmPassword", "")
                if not new_password:
                    raise ValueError("New password is required")
                if new_password != confirm_password:
                    raise ValueError("New passwords did not match")
                change_own_password(user, current_password, new_password)
                self.send_json({"ok": True})
            except ValueError as exc:
                self.send_json({"error": str(exc)}, 400)
            except (json.JSONDecodeError, UnicodeDecodeError):
                self.send_json({"error": "Invalid JSON body"}, 400)
            return

        if path == "/api/admin/reset-password":
            user = self.require_auth(api=True)
            if user is None:
                return
            if not user.get("isAdmin"):
                self.send_json({"error": "Administrator access required"}, 403)
                return
            try:
                payload = self.read_json_body()
                target = admin_reset_password(
                    user,
                    payload.get("username", ""),
                    payload.get("newPassword", ""),
                )
                self.send_json(
                    {
                        "ok": True,
                        "message": f"{target['displayName']} must change their password after the next login.",
                        "user": target,
                    }
                )
            except ValueError as exc:
                self.send_json({"error": str(exc)}, 400)
            except PermissionError as exc:
                self.send_json({"error": str(exc)}, 403)
            except (json.JSONDecodeError, UnicodeDecodeError):
                self.send_json({"error": "Invalid JSON body"}, 400)
            return

        self.send_error(404)


def main():
    ensure_app_state_db()
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(sys.argv[1] if len(sys.argv) > 1 else os.environ.get("PORT", "8765"))
    server = ThreadingHTTPServer((host, port), Handler)
    print(f"CDX CRM UI running on {host}:{port}")
    print(f"Database path: {DB_PATH}")
    print(f"Attachment archive path: {BACKUP_DIR}")
    print(f"App state path: {APP_STATE_DB_PATH}")
    print("Press Ctrl+C to stop.")
    server.serve_forever()


if __name__ == "__main__":
    main()
